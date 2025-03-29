import streamlit as st
st.set_page_config(
    page_title="Integrated Stock Sentiment App",
    page_icon="icon.png",  # Buraya ikonun dosya adını yaz
    layout="wide"
)
import feedparser
import requests
import pandas as pd
import plotly.graph_objs as go
import time
from transformers import pipeline
from streamlit.components.v1 import html as st_html
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import unicodedata
import os
import yfinance as yf
from pandas_datareader import data as pdr
import numpy as np
from arch import arch_model


if "investment_decision" not in st.session_state:
    st.session_state["investment_decision"] = None


# -----------------------------
# Alpha Vantage & Model Config
# -----------------------------
ALPHA_VANTAGE_API_KEYS = [
    "1JBBBFINKKCNWNQO.",
    "EJU22Q5IIIASLS3V.",
]
ALPHA_VANTAGE_API_URL = 'https://www.alphavantage.co/query'
MODEL_NAME = "ProsusAI/finbert"

# -----------------------------
# Scrolling Ticker Functions
# -----------------------------
stock_tickers = [
    'ABT', 'AMP', 'T', 'AZO', 'AXON', 'BAC', 'BRO', 'CZR', 'SCHW', 'DTE', 
    'GNRC', 'HD', 'MNST', 'NVDA', 'OKE', 'PNR', 'PGR', 'SYK', 'TRMB', 'WMT', 
    'TSLA', 'ALB', 'WYNN', 'PLTR', 'TRGP', 'COP', 'HES', 'PAYC', 'KMX', 
    'MOS', 'UAL', 'ON', 'CF', 'SPG', 'VLO', 'RCL.GF', 'BKR', 'DELL', 'TPR', 
    'SYF', 'LW', 'AIG', 'FANG', 'DFS', 'WDC', 'AMD', 'EOG', 'BA', 'DAL', 
    'ANET', 'FCX', 'MPC', 'VTR', 'AMZN', 'PYPL', 'EQT'
]



def get_stock_prices(tickers):
    prices = {}
    for i, ticker in enumerate(tickers):
        try:
            stock = yf.download(ticker, period="1d")
            if not stock.empty:
                prices[ticker] = stock['Close'].iloc[-1]  # Burada Series yerine float alıyoruz
        except Exception as e:
            print(f"Error fetching {ticker}: {e}")
        
        # API limitine takılmamak için her 5 istekte bir bekleyelim
        if i % 5 == 0:
            time.sleep(2)  # 2 saniye bekleme
    return prices


def get_stock_data(ticker):
    """
    Hisse senedi fiyatlarını çekerek Momentum, RSI ve Sharpe Ratio hesaplar.
    """
    try:
        stock = yf.download(ticker, period="6mo")  # Son 6 ayın verisini çekiyoruz
        if stock.empty:
            return None

        # Momentum (10 günlük)
        stock['Momentum'] = stock['Close'] - stock['Close'].shift(12)

        # RSI Hesaplama (14 günlük) - Makale ile Tam Uyumlu

        period = 14  # RSI periyodu

        # Günlük fiyat değişimlerini hesapla
        stock['delta'] = stock['Close'].diff()

        # Kazanç ve kayıpları ayır
        stock['gain'] = stock['delta'].where(stock['delta'] > 0, 0)
        stock['loss'] = -stock['delta'].where(stock['delta'] < 0, 0)

        # Ortalama kazanç ve kayıp hesapla (Wilder'ın üstel hareketli ortalaması)
        stock['avg_gain'] = stock['gain'].ewm(alpha=1/period, adjust=False).mean()
        stock['avg_loss'] = stock['loss'].ewm(alpha=1/period, adjust=False).mean()

        # RS (Relative Strength) hesapla
        epsilon = 1e-10  # Sıfıra bölme hatasını önlemek için küçük bir değer ekle
        stock['rs'] = stock['avg_gain'] / (stock['avg_loss'] + epsilon)

        # RSI hesaplama
        stock['RSI'] = 100 - (100 / (1 + stock['rs']))


        # Getiri (Return) Hesaplama
        stock['Return'] = stock['Close'].pct_change().dropna()

        # 📌 GARCH(1,1) Modeli ile Volatilite Tahmini
        model = arch_model(stock['Return'].dropna(), vol='Garch', p=1, q=1)
        model_fit = model.fit(disp="off")

        # GARCH Modeli ile Tahmin Edilen Volatilite
        stock['GARCH_Volatility'] = model_fit.conditional_volatility

        # GARCH tabanlı Sharpe Ratio Hesaplama
        risk_free_rate = 0.02  # %2 risksiz faiz oranı
        expected_return = stock['Return'].mean()  # Ortalama getiri
        volatility = stock['GARCH_Volatility'].mean()  # GARCH ile tahmin edilen volatilite

        # Eğer volatilite 0 ise bölme hatası olmaması için kontrol edelim
        sharpe_ratio = (expected_return - risk_free_rate) / volatility if volatility > 0 else np.nan

        # Önceki günün RSI değerini hesapla
        stock['prev_RSI'] = stock['RSI'].shift(1)

        latest_data = stock.iloc[-1]  # En son günün verisini al
        prev_rsi_value = stock['prev_RSI'].iloc[-1]  # Önceki günün RSI değerini al

        return {
            "Price": latest_data["Close"],
            "Momentum": latest_data.get("Momentum", None),
            "RSI": latest_data.get("RSI", None),
            "prev_RSI": prev_rsi_value,  # Önceki günün RSI değerini ekledik
            "Sharpe Ratio": sharpe_ratio
        }
    except Exception as e:
        print(f"Error fetching stock data for {ticker}: {e}")
        return None


def display_scrolling_ticker(prices):
    """
    Generates and displays an HTML-based ticker using CSS animations.
    """
    ticker_text = ' | '.join([f"{ticker}: ${price.iloc[-1]:.2f}" if isinstance(price, pd.Series) else f"{ticker}: ${price:.2f}" for ticker, price in prices.items()])
    html_code = f"""
    <div class="ticker-container">
      <div class="ticker-text">
        {ticker_text}
      </div>
    </div>
    """
    st.markdown(html_code, unsafe_allow_html=True)

# -----------------------------
# Stock & Sentiment Functions
# -----------------------------


def analyze_sentiment(text):
    pipe = pipeline("text-classification", model=MODEL_NAME)
    sentiment = pipe(text)[0]
    return sentiment["label"], sentiment["score"]

def fetch_stock_data(symbol):
    """
    Fetches intraday stock data from Alpha Vantage with multiple API keys.
    """
    for api_key in ALPHA_VANTAGE_API_KEYS:
        print(f"🔄 Trying API Key: {api_key}")  # Hangi API kullanılıyor gör
        params = {
            'function': 'TIME_SERIES_INTRADAY',
            'symbol': symbol,
            'interval': '5min',
            'apikey': api_key
        }
        response = requests.get(ALPHA_VANTAGE_API_URL, params=params)
        print(f"🔍 API Response Status Code: {response.status_code}")  # HTTP yanıt kodunu kontrol et
        data = response.json()
        print(f"🔍 API Response Data: {data}")  # JSON dönen cevabı göster

        if 'Time Series (5min)' in data:
            df = pd.DataFrame.from_dict(data['Time Series (5min)'], orient='index', dtype=float)
            df = df.rename(columns={
                '1. open': 'Open',
                '2. high': 'High',
                '3. low': 'Low',
                '4. close': 'Close',
                '5. volume': 'Volume'
            })
            df.index = pd.to_datetime(df.index)
            df.sort_index(inplace=True)
            return df
        else:
            print(f"❌ API Key {api_key} failed. Trying next key...")

    st.error("❌ All API keys failed. Please check your API limits or try again later.")
    return None

def plot_stock_data(df, symbol):
    """
    Plots the stock's closing price over time using Plotly.
    """
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df.index, y=df['Close'], mode='lines', name='Close Price'))
    fig.update_layout(
        title=f'Real-Time Stock Price for {symbol}',
        xaxis_title='Time',
        yaxis_title='Price (USD)',
        xaxis_rangeslider_visible=False,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font_color='#F5F5F5'
    )
    st.plotly_chart(fig, use_container_width=True)

def get_sp500_tickers():
    """Fetches a static list of S&P 500 tickers."""
    tickers = [
    'ABT', 'AMP', 'T', 'AZO', 'AXON', 'BAC', 'BRO', 'CZR', 'SCHW', 'DTE', 
    'GNRC', 'HD', 'MNST', 'NVDA', 'OKE', 'PNR', 'PGR', 'SYK', 'TRMB', 'WMT', 
    'TSLA', 'ALB', 'WYNN', 'PLTR', 'TRGP', 'COP', 'HES', 'PAYC', 'KMX', 
    'MOS', 'UAL', 'ON', 'CF', 'SPG', 'VLO', 'RCL.GF', 'BKR', 'DELL', 'TPR', 
    'SYF', 'LW', 'AIG', 'FANG', 'DFS', 'WDC', 'AMD', 'EOG', 'BA', 'DAL', 
    'ANET', 'FCX', 'MPC', 'VTR', 'AMZN', 'PYPL', 'EQT'
    ]
    return tickers

# -----------------------------
# Utility Functions for Text & Excel
# -----------------------------
def clean_text(text):
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")

def shorten_title(title, length=80):
    return title if len(title) <= length else title[:length] + "..."


def decision_making(rsi, prev_rsi, momentum, sharpe, sentiment_score):
    buy_signals = 0
    sell_signals = 0

    # 📌 **Sentiment ağırlığı artırıldı (2x)**
    sentiment_weight = 2  



    # RSI Kararı (Makale ile uyumlu)
    if rsi > 30 and prev_rsi <= 30:  
        buy_signals += 1  

    elif rsi > 70 and prev_rsi <= 70:  
        sell_signals += 1  

    # Momentum Kararı (Ağırlık: 1)
    if momentum > 0:
        buy_signals += 1
    elif momentum < 0:
        sell_signals += 1

    # Sharpe Ratio Kararı (Ağırlık: 1)
    if sharpe > 1:
        buy_signals += 1
    elif sharpe < 0:
        sell_signals += 1

    # **Sentiment Kararı (Ağırlık: 2)**
    if sentiment_score > 0.5:
        buy_signals += sentiment_weight
    elif sentiment_score < -0.5:
        sell_signals += sentiment_weight

    # 📌 **Güncellenmiş Karar Mekanizması**
    if buy_signals > sell_signals:
        return "BUY"
    elif sell_signals > buy_signals:
        return "SELL"
    else:
        return "HOLD"


def create_excel_with_formatting(ticker, articles):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{ticker[:30]} Analysis"  # Truncate title if necessary

    # Title above the columns
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Sentiment Analysis for {ticker[:30]}"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Two empty rows after the title
    ws.append([])
    ws.append([])

    # Column headers
    headers = ["Title", "Sentiment", "Published", "Summary", "Link"]
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data with sentiment-based row coloring
    sorted_articles = sorted(articles, key=lambda x: (x["sentiment"], x["published"]), reverse=True)
    for article in sorted_articles:
        sentiment = article["sentiment"]
        color = (
            "d4edda" if sentiment == "positive" else "f8d7da" if sentiment == "negative" else "fff3cd"
        )
        ws.append([
            article["title"],
            sentiment.capitalize(),
            article["published"],
            article["summary"],
            article["link"][:50] + "..."  # Truncate the link for display
        ])
        row_num = ws.max_row
        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(size=12)

    # Add sentiment count summary
    ws.append([])
    ws.append([])
    ws.append(["Sentiment", "Count"])
    ws.append(["Positive", sum(1 for article in articles if article["sentiment"] == "positive")])
    ws.append(["Neutral", sum(1 for article in articles if article["sentiment"] == "neutral")])
    ws.append(["Negative", sum(1 for article in articles if article["sentiment"] == "negative")])
    for row in range(ws.max_row - 3, ws.max_row + 1):
        for col in range(1, 3):
            ws.cell(row=row, column=col).font = Font(bold=(row == ws.max_row - 3), size=12)

    # Set uniform column widths
    for col_letter in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 40.0

    return wb

# -----------------------------
# Main Streamlit App
# -----------------------------
def main():
    # Set wide layout and custom page title





    # --- Inject Custom CSS for Modern Look & Article Boxes ---
# Inject Custom CSS for the updated design
# Inject Custom CSS for the updated design
# Inject Custom CSS for the updated design with white overall background
 # Inject Custom CSS for white background and black text
    st.markdown(
        """
        <style>
        /* Import a modern font from Google Fonts */
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
        html, body {
            background-color: #ffffff !important; /* Beyaz arka plan */
            font-family: 'Poppins', sans-serif;
            color: #000000 !important; /* Siyah yazılar */
        }
        .block-container {
            background-color: transparent !important;
            padding: 2rem !important;
        }
        
        h1, h2, h3, h4, h5 {
            color: #ff7f00 !important; /* Turuncu renk başlıklar için */
            margin-top: 1rem;
        }
        .ticker-container {
            background-image: linear-gradient(to right, #4caf50, #8bc34a); /* Yeşil gradyanı */
            color: #fff;
            overflow: hidden;
            white-space: nowrap;
            box-sizing: border-box;
            padding: 10px 0;
            margin-bottom: 20px;
            border-radius: 10px;
        }
        .ticker-text {
            display: inline-block;
            padding-left: 100%;
            font-size: 1.2rem;
            font-weight: bold;
            animation: ticker 30s linear infinite;
        }
        @keyframes ticker {
            0%   { transform: translate3d(0, 0, 0); }
            100% { transform: translate3d(-100%, 0, 0); }
        }
        .stButton>button {
            background-color: #ff7f00; /* Turuncu buton */
            color: #ffffff;
            border-radius: 8px;
            padding: 0.6em 1.2em;
            font-weight: 600;
            border: none;
            transition: background-color 0.2s ease-in-out, color 0.2s ease-in-out;
        }
        .stButton>button:hover {
            background-color: #8bc34a; /* Hoverda yeşil */
            color: #fff;
            cursor: pointer;
        }
        .content-card {
            background-color: #ffffff; /* İçerik kartları için beyaz arka plan */
            border: 1px solid #e0e0e0; /* Hafif border */
            border-radius: 10px;
            padding: 1.5rem;
            margin-bottom: 2rem;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1); /* Hafif gölge */
        }
        .sentiment-box {
            background-color: #ffffff;
            border-radius: 8px;
            padding: 1rem;
            margin-top: 1rem;
            color: #ff7f00; /* Turuncu metin */
            font-weight: 600;
        }
        /* Makale kutuları için stil */
        .article-box {
            padding: 10px;
            margin-bottom: 15px;
            border-radius: 5px;
            border: 2px solid;
        }
        .positive {
            border-color: #4caf50; /* Pozitif için yeşil */
        }
        .neutral {
            border-color: #ffeb3b; /* Nötr için sarı */
        }
        .negative {
            border-color: #dc3545; /* Negatif için kırmızı */
        }
        .article-title {
            font-size: 16px;
            font-weight: bold;
        }
        .sentiment-arrow {
            font-size: 14px;
            font-weight: bold;
            display: inline-block;
            margin-right: 5px;
        }
        /* Grafikler için stil */
        .plotly-graph-div {
            background-color: #ffffff !important; /* Grafik arka planı beyaz */
        }
        </style>
        """,
        unsafe_allow_html=True
    )




    st.title("Daily Stock Analysis - InvestingAI-6")

    # --- Ticker Section ---
    st.write("### Live Stock Prices")
    with st.container():
        prices = get_stock_prices(stock_tickers)
        display_scrolling_ticker(prices)

    st.write("---")

    # --- Sentiment Analysis Controls ---
    st.write("### Select the stock")
    pipe = pipeline("text-classification", model=MODEL_NAME)
    use_sp500 = st.checkbox("Choose ticker from S&P 500 list", value=True)
    if use_sp500:
        sp500_list = get_sp500_tickers()
        default_index = sp500_list.index("META") if "META" in sp500_list else 0
        ticker = st.selectbox("Select Stock Ticker", sp500_list, index=default_index)
    else:
        ticker = st.text_input("Enter Stock Ticker (e.g., META):", "META")
    keyword = st.text_input("Enter Keyword for Filtering Articles:", ticker.lower())
    
    # Option to export results as Excel
    export_excel = st.checkbox("Export Results to Excel", value=False)


    # --- Analysis Button ---
    if st.button("Analyze"):
        st.write("## Results")

        # 1. Load selected sentiment analysis model
        pipe = pipeline("text-classification", model=MODEL_NAME)


        # 2. Fetch stock data & display chart
        with st.spinner(f"Fetching stock data for {ticker}..."):
            stock_data = fetch_stock_data(ticker)
            if stock_data is not None:
                with st.container():
                    st.write(f"### Intraday Stock Chart for {ticker}")
                    plot_stock_data(stock_data, ticker)

        # Ek metrikleri alıyoruz (Momentum, RSI, Sharpe Ratio)
        stock_metrics = get_stock_data(ticker)
        if stock_metrics:
            st.write(f"### {ticker} Metrics")

            # 3 Kutu yan yana göstermek için
            col1, col2, col3 = st.columns(3)

            with col1:
                st.metric(label="RSI", value=f"{float(stock_metrics['RSI']):.2f}")

            with col2:
                st.metric(label="Momentum", value=f"{float(stock_metrics['Momentum']):.2f}")

            with col3:
                st.metric(label="Sharpe Ratio", value=f"{float(stock_metrics['Sharpe Ratio']):.2f}")


        else:
            st.error("Stock metrics could not be calculated.")

        # 3. Fetch RSS feed & run sentiment analysis on news articles
        rss_url = f"https://finance.yahoo.com/rss/headline?s={ticker}"
        with st.spinner("Fetching and analyzing news articles..."):
            feed = feedparser.parse(rss_url)
            articles_list = []
            total_score = 0
            num_articles = 0
            start_time = time.time()

            for entry in feed.entries:
                # Filter based on keyword in title or summary
                if keyword.lower() not in entry.summary.lower() and keyword.lower() not in entry.title.lower():
                    continue

                sentiment_label, sentiment_score = analyze_sentiment(entry.summary)


                # Map sentiment label to one of positive/negative/neutral
                if "pos" in sentiment_label.lower():
                    mapped_label = "positive"
                    total_score += sentiment_score
                    num_articles += 1
                elif "neg" in sentiment_label.lower():
                    mapped_label = "negative"
                    total_score -= sentiment_score
                    num_articles += 1
                else:
                    mapped_label = "neutral"


                articles_list.append({
                    "title": entry.title,
                    "published": entry.published,
                    "summary": entry.summary,
                    "link": entry.link,
                    "sentiment": mapped_label,
                    "score": sentiment_score

                })

            elapsed_time = time.time() - start_time

        # 4. Overall sentiment summary
        if num_articles > 0:
            if total_score > 0.15:
                overall_sentiment = "Positive"
            elif total_score < -0.15:
                overall_sentiment = "Negative"
            else:
                overall_sentiment = "Neutral"

            st.markdown(f"""
            <div class="sentiment-box">
            <h4>Overall Sentiment: {overall_sentiment}</h4>
            <p>Aggregated Score: {total_score:.2f}</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.error("No articles found matching the keyword.")


# -----------------------------
# Investment Decision Button
# -----------------------------

        # 📌 Yatırım Kararını Hafızada Tut
        if "investment_decision" not in st.session_state:
            st.session_state["investment_decision"] = None

        # ✅ **Analyze Butonuna Tıklanınca Otomatik Karar Hesapla**
        if stock_metrics:
            st.session_state["investment_decision"] = decision_making(
                rsi=float(stock_metrics["RSI"]),
                prev_rsi=float(stock_metrics["prev_RSI"]),  # Önceki RSI değerini ekledik
                momentum=float(stock_metrics["Momentum"]),
                sharpe=float(stock_metrics["Sharpe Ratio"]),
                sentiment_score=float(total_score)
            )

        # 📌 **Tek bir yatırım kararı başlığı**
        st.write("### Investment Decision")

        # 📌 **Tek bir yatırım kararı kutusu**
        if st.session_state["investment_decision"]:
            st.markdown(f"""
            <div style="padding: 20px; border-radius: 10px; background-color: #1e1e1e; color: white; 
                        font-weight: bold; text-align: center; font-size: 20px; border: 2px solid #ff9900;">
            Investment Decision: {st.session_state["investment_decision"]}
            </div>
            """, unsafe_allow_html=True)


        # 5. Sentiment distribution chart using Plotly in a centered column
        positive_count = sum(1 for a in articles_list if a["sentiment"] == "positive")
        neutral_count = sum(1 for a in articles_list if a["sentiment"] == "neutral")
        negative_count = sum(1 for a in articles_list if a["sentiment"] == "negative")
        labels = ["Positive", "Neutral", "Negative"]
        values = [positive_count, neutral_count, negative_count]
        fig = go.Figure(data=[go.Bar(
            x=labels,
            y=values,
            marker_color=["#28a745", "#ffc107", "#dc3545"]
        )])
        fig.update_layout(
            title=f"Sentiment Distribution for {ticker}",
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font_color="#F5F5F5"
        )

        # Use columns to center the chart
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.plotly_chart(fig, use_container_width=True)



        # 7. Display articles with colored sentiment boxes
        st.write(f"### Articles ({len(articles_list)})")
        for article in articles_list:
            sentiment_class = "positive" if article["sentiment"] == "positive" else "negative" if article["sentiment"] == "negative" else "neutral"
            st.markdown(
                f"""
                <div class="article-box {sentiment_class}">
                <div class="article-title">{shorten_title(clean_text(article['title']))}</div>
                <strong>Published:</strong> {article.get('published', 'N/A')}<br>
                <strong>Summary:</strong> {clean_text(article['summary'])}<br>
                <span class="sentiment-arrow">&rarr;</span> {article['sentiment'].capitalize()} ({article['score']:.2f})<br>
                <a href="{article['link']}" target="_blank">Read Full Article</a>
                </div>
                """, unsafe_allow_html=True
            )

        # 8. Excel export functionality
        if export_excel and articles_list:
            wb = create_excel_with_formatting(ticker, articles_list)
            excel_data = BytesIO()
            wb.save(excel_data)
            excel_data.seek(0)
            st.download_button(
                label="Download Analysis (Excel)",
                data=excel_data,
                file_name=f"{ticker}_sentiment_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # 9. Display performance stats
        st.write("---")
        st.write(f"**Inference Time (seconds):** {elapsed_time:.2f}")

if __name__ == "__main__":
    main()
